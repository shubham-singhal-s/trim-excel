from openpyxl import load_workbook
import csv

file = open('output.csv', 'w', newline='')
writer = csv.writer(file)
workbook = load_workbook(filename="input.xlsx")
sheet = workbook.active
for row in sheet.iter_rows(values_only=True):
    arr = []
    for cell in row:
        arr.append(cell.strip())
    writer.writerow(arr)
